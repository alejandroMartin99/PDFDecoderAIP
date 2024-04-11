from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def save_format_frontend(output_path, dataframes, sheet_names):
    # Crear un nuevo libro de Excel
    wb = Workbook()

    for df, sheet_name in zip(dataframes, sheet_names):
        if df.empty:
            continue

        else:

            ws = wb.create_sheet(title=sheet_name)

            # Configurar el ancho de las columnas
            column_width=16
            for col_num, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = column_width

            # Escribir el encabezado
            for col_num, value in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=value).font = Font(bold=True)
                ws.cell(row=1, column=col_num).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

            # Escribir los datos
            for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

                    # Aplicar formato de contraste a las celdas
                    if row_num % 2 == 0:
                        ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            max_column_letter = get_column_letter(ws.max_column)
            ref = f"A1:{max_column_letter}{ws.max_row}"
            table = Table(displayName=sheet_name, ref=ref)

            # table = Table(displayName=sheet_name, ref=f"A1:{chr(64 + ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(
                name="TableStyleLight1", 
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            
    # Eliminar la hoja predeterminada que se crea al inicio
    wb.remove(wb.active)

    # Guardar el libro de trabajo en el archivo Excel
    wb.save(output_path)

